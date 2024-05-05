import openpyxl
from xls2xlsx import XLS2XLSX
import os
del_flag = False

class Pos :
    def __init__(self):
        self.column = 0
        self.row = 0
class Cells :
    def __init__(self):
        self.start = Pos()
        self.end = Pos()
        rowrange = 0
        columnrange = 0

def get_merge_range(sheet):
    merged_ranges = []
    for i in sheet.merged_cells.ranges:
        mergecell = Cells()
        mergecell.start.column = i.min_col
        mergecell.start.row = i.min_row
        mergecell.end.column = i.max_col
        mergecell.end.row = i.max_row
        mergecell.rowrange = i.size['rows']
        mergecell.columnrange = i.size['columns']
        merged_ranges.append(mergecell)
    return merged_ranges

def getfile(filepath):
    filename = "Unknown"
    if filepath.find("\\") != -1:
        filename = filepath.split("\\")[-1]
    elif filepath.find("/") != -1:
        filename = filepath.split("/")[-1]
    else:
        filename = filepath
    if(filename.find('"') != -1):
        filename = filename[:-1]
    return filename

filepath = input("File Path: ")
filename = getfile(filepath)
if(filename[-3:]=="xls"):
    tmpfile = "tmp.xlsx"
    x2x = XLS2XLSX(filename)
    x2x.to_xlsx(tmpfile)
    filepath = "H:\\resources\\Programs\\tmp.xlsx"
    del_flag = True
output = open(filename + ".html", "w",encoding = "utf8")
document = openpyxl.load_workbook(filepath,data_only=True)
worksheet = document.active
ChangeFlag = False
# maxcols = 0
# cols = 0
output.write('<!DOCTYPE html>\n<html>\n<head>\n')
output.write('<style type="text/css">\ntable {table-layout: fixed;border-collapse: collapse;border: 5px solid grey;text-align: center;}\nth, td {padding: 2px;}\ncaption{font-weight: bold;}\ntd{border: 2px grey solid}\n.y{background-color: rgb(255, 232, 156)}\n.g{background-color: rgb(195, 255, 242)}\n</style>\n')
output.write('</head>\n<body>\n<table border>\n<caption>' + filename + '</caption>')
merged_ranges = get_merge_range(worksheet)
for row in worksheet.iter_rows():
    output_string = "<tr>\n"
    # output.write("<tr>\n")
    cols = 0
    for cell in row:
        written = False
        for mergecell in merged_ranges:
            if cell.column >= mergecell.start.column and cell.row >= mergecell.start.row and cell.row <= mergecell.end.row and cell.column <= mergecell.end.column:
                written = True
                if cell.column == mergecell.start.column and cell.row == mergecell.start.row:
                    if ChangeFlag:
                        output_string += '<td class="y" rowspan = "' + str(mergecell.rowrange) + '" colspan = "' + str(mergecell.columnrange) + '">' + str(cell.value if cell.value != None else "") + '</td>\n'
                    else:
                        output_string += '<td class="g" rowspan = "' + str(mergecell.rowrange) + '" colspan = "' + str(mergecell.columnrange) + '">' + str(cell.value if cell.value != None else "") + '</td>\n'
                    # output.write('<td rowspan = "' + str(mergecell.rowrange) + '" colspan = "' + str(mergecell.columnrange) + '">' + str(cell.value) + '</td>\n')
        if not written:
            if ChangeFlag:
                output_string += '<td class="y">' + str(cell.value if cell.value != None else "") + '</td>\n'
            else:
                output_string += '<td class="g">' + str(cell.value if cell.value != None else "") + '</td>\n'
            # output.write("<td>" + str(cell.value) + "</td>\n")
        cols += 1
    # if cols > maxcols:
    #     maxcols = cols
    output.write(output_string + '</tr>\n')
    ChangeFlag = not ChangeFlag
output.write("</table>\n</body>\n</html>")
try:
    os.remove("H:\\resources\\Programs\\tmp.xlsx")
except:
    None