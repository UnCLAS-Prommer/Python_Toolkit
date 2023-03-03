import openpyxl

class Pos :
    def __init__(self):
        self.column = 0
        self.row = 0
class Cells :
    def __init__(self):
        self.start = Pos()
        self.end = Pos()
        rowrange = 0
        columnrange = 0

def get_merge_range(sheet):
    merged_ranges = []
    for i in sheet.merged_cells.ranges:
        mergecell = Cells()
        mergecell.start.column = i.min_col
        mergecell.start.row = i.min_row
        mergecell.end.column = i.max_col
        mergecell.end.row = i.max_row
        mergecell.rowrange = i.size['rows']
        mergecell.columnrange = i.size['columns']
        merged_ranges.append(mergecell)
    return merged_ranges

filepath = input("File Path: ")
output = open("output.html", "w",encoding = "utf8")
filename = "Unknown"
if filepath.find("\\") != -1:
    filename = filepath.split("\\")[-1]
elif filepath.find("/") != -1:
    filename = filepath.split("/")[-1]
else:
    filename = filepath
document = openpyxl.load_workbook(filepath,data_only=True)
worksheet = document.active
maxcols = 0
cols = 0
output.write('<!DOCTYPE html>\n<html>\n<head>\n')
output.write('<style type="text/css">\ntable {table-layout: fixed;width: 100%;border-collapse: collapse;border: 3px solid purple;text-align: center;}\nth, td {padding: 2px;}\ncaption{font-weight: bold;}\n</style>\n')
output.write('</head>\n<body>\n')
output.write('<table border="2">\n<caption>' + filename + '</caption>')
merged_ranges = get_merge_range(worksheet)
for row in worksheet.iter_rows():
    output.write("<tr>\n")
    cols = 0
    for cell in row:
        written = False
        for mergecell in merged_ranges:
            if cell.column >= mergecell.start.column and cell.row >= mergecell.start.row and cell.row <= mergecell.end.row and cell.column <= mergecell.end.column:
                written = True
                if cell.column == mergecell.start.column and cell.row == mergecell.start.row:
                    output.write('<td rowspan = "' + str(mergecell.rowrange) + '" colspan = "' + str(mergecell.columnrange) + '">' + str(cell.value) + '</td>\n')
        if not written:
            output.write("<td>" + str(cell.value) + "</td>\n")
        cols += 1
    if cols > maxcols:
        maxcols = cols
    output.write("</tr>\n")
output.write("</table>\n</body>\n</html>")